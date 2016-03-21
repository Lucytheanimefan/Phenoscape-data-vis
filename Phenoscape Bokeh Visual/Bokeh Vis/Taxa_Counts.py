'''
Created on Mar 11, 2016

@author: lucy
'''
import openpyxl
import json
from urllib2 import urlopen, Request
from collections import OrderedDict


# Open spreadsheet of VTO and Term IRIs
data=openpyxl.load_workbook('VTO.xlsx')
sheet=data.get_sheet_by_name('Worksheet')
print sheet['A1'].value

# array of total annotated taxa counts 
labels=[]
num=[]
count={}
for i in range(2,10):
    print(sheet.cell(row=i, column=1).value)
    cellLabel=sheet.cell(row=i, column=2).value
    cellValue=sheet.cell(row=i, column=1).value
    url='http://kb.phenoscape.org/api/taxon/annotated_taxa_count?in_taxon='+cellValue
    taxaCount=json.loads(urlopen(url).read())
    labels.append(cellLabel)
    num.append(taxaCount['total'])

count['label']=labels
count['taxaCount']=num
    
print count

# method finds url depending on name of taxa/term
def get_url(nameOfTaxa):
    for i in range(2,sheet.get_highest_row()-1):
        cellLabel=sheet.cell(row=i, column=2).value
        if cellLabel==nameOfTaxa:
            cellValue=sheet.cell(row=i, column=1).value
            print cellValue
            return cellValue
        
get_url('Hominoidea')

